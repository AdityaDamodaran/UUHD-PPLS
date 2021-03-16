"""
SPDX-FileCopyrightText: 2021 University of Luxembourg
SPDX-License-Identifier: GPL-3.0-or-later
SPDXVersion: SPDX-2.2

Authors: 
       Aditya Damodaran, aditya.damodaran@uni.lu
       Alfredo Rial, alfredo.rial@uni.lu
"""

import argparse
import time
import random
import sys
from pathlib import Path
from collections import namedtuple

from charm.toolbox.pairinggroup import PairingGroup, ZR, G1, G2
from openpyxl import load_workbook, Workbook
from texttable import Texttable

from uuhd.jsonobjects import (
    ZKWitness,
    ZKInstance,
    SubWitnessRecord,
    SubInstanceRecord,
    dict_from_class,
)
from uuhd.functionalities import (
    FNYM,
    FCRS,
    FREG,
    FZK,
    FZK_RD,
    FZK_PR3,
    WeakReference,
)
from uuhd.primitives import (
    VectorCommitment,
    StructurePreservingSignature,
    PedersenCommitment,
)
from uuhd.measurement_util import get_real_size

# Note: The variable names used here may not reflect the actual names
# used in our paper because we have renamed them for PEP-8 compliance.
# However, dictionary keys and messages use names from the paper
# for brevity.


# NTs
CommittedRecord = namedtuple(
    "CommittedRecord", "i vr ccom_i copen_i ccom_ri copen_ri"
)
WitnessRecord = namedtuple(
    "WitnessRecord", "i vr copen_i copen_ri witness sig_i"
)

InstanceRecord = namedtuple("InstanceRecord", "ccom_i ccom_ri")


def draw_table(headings, data):
    """Prints an ASCII table to console."""
    table = Texttable()
    table.add_rows([headings, data])
    print(table.draw())


def setup(nsize):
    """Sets up the CRS (public parameters for a Vector Commitment with
    a database of size 'nsize', and parameters for the
    Pedersen commitment scheme)."""
    par = vector_commitment.setup(nsize)
    par_c = pedersen_commitment.setup()
    f_crs.set(par, par_c)


class Updater:
    """Party U (Updater)."""

    def __init__(self):
        pass

    l_store = []
    #   'sid', 'P', 'flag=1', 'com1'
    #   'sid', 'P', 'flag=0', 'vcom', 'com', 'coms'

    l_par = []
    #   'sid', 'par', 'par_c', 'sps_public_key', 'sps_secret_key'

    def is_unique_pseudonym(self, pseudonym):
        """Checks if pseudonym p has been seen before."""
        for item in self.l_store:
            if pseudonym == item["p"]:
                return 0
        return 1

    def get_record_for_pseudonym(self, pseudonym):
        """Returns the corresponding record in l_store
        for pseudonym = p."""
        for item in self.l_store:
            if pseudonym == item["p"]:
                return item
        return 0

    def message_in(self, m, pseudonym):
        """Handles messages from other parties/functionalities to U."""
        if m["message"] == "com1":
            if not self.is_unique_pseudonym(pseudonym):
                print("Abort: (Updater) p is not unique. (com1)")
                exit()
            flag = 1
            self.l_store.append(
                {"sid": sid, "p": pseudonym, "flag": flag, "com1": m["com1"]}
            )
        elif m["message"] == "repcom":
            record = self.get_record_for_pseudonym(pseudonym)
            if (
                pedersen_commitment.verify(
                    self.l_par[0]["par_c"], record["com"], m["s1"], m["open1"]
                )
                == 1
            ):
                record["com1"] = m["com1"]
            else:
                print("Abort: (Updater) Commitment opening failed. (repcom)")
                exit()
            if args.verbose:
                com_str = "; coms = " + str(record["coms"])
            else:
                com_str = "; coms"
            print(
                "Updater: hd.read.end; sid = "
                + str(record["sid"])
                + "; p = "
                + str(record["p"])
                + "; flag = "
                + str(record["flag"])
                + com_str
            )
        else:
            print(
                "Abort: (Updater) Unrecognised message: " + str(m["message"])
            )
            exit()

    def proof_in(self, instance, pseudonym):
        """Handles proof related messages from other FZKs to U."""
        if not self.is_unique_pseudonym(pseudonym):
            print("Abort: (Updater) p is not unique. (proof)")
            exit()
        if (
            instance["sid"] != self.l_par[0]["sid"]
            or instance["par"]["par_g"] != self.l_par[0]["par"]["par_g"]
            or instance["par"]["par_h"] != self.l_par[0]["par"]["par_h"]
            or instance["par_c"]["g"] != self.l_par[0]["par_c"]["g"]
            or instance["par_c"]["h"] != self.l_par[0]["par_c"]["h"]
        ):
            print("Abort: (Updater) Inconsistent instance vars. (proof)")
            exit()
        self.l_store.append(
            {
                "sid": sid,
                "p": pseudonym,
                "flag": 0,
                "vcom": instance["vcomd"],
                "com": instance["comd"],
                "coms": instance["ins_i"],
            }
        )
        f_nym.reply({"message": "open", "com": instance["comd"]}, pseudonym)

    def update(self, sid, pseudonym, values):
        """Update interface."""
        record = self.get_record_for_pseudonym(pseudonym)

        if record == 0:
            print("Abort: (Updater) No records found for p. (update)")
            exit()

        if len(self.l_par) == 0:
            t_setup_start = time.time()
            par, par_c = f_crs.get()

            par_sig = [0]
            (
                sps_public_key,
                sps_secret_key,
            ) = structure_preserving_signature.generate_keys(
                {"group": par["group"], "g": par["g"], "h": par["h"]}, 2, 1
            )

            for i in range(1, len(par["par_h"])):
                sps_r, sps_s, sps_t = structure_preserving_signature.sign(
                    sps_secret_key,
                    [0, par["g"] ** i, par["g"] ** sid, par["par_h"][i]],
                )
                par_sig.append({"R": sps_r, "S": sps_s, "T": sps_t})

            f_reg.register(
                {"sps_public_key": sps_public_key, "par_sig": par_sig}
            )

            self.l_par.append(
                {
                    "sid": sid,
                    "par": par,
                    "par_c": par_c,
                    "sps_public_key": sps_public_key,
                    "sps_secret_key": sps_secret_key,
                }
            )

            global t_setup
            t_setup = str(time.time() - t_setup_start)

        x = [0]

        for value_item in values:
            x.append(value_item)

        vcom = vector_commitment.commit(self.l_par[0]["par"], x, 0)
        s2 = self.l_par[0]["par"]["group"].random(ZR)
        com2 = pedersen_commitment.commit_0(self.l_par[0]["par_c"], s2)
        com = record["com1"] * com2["com"]

        if record["flag"] == 1:

            sps_r, sps_s, sps_t = structure_preserving_signature.sign(
                self.l_par[0]["sps_secret_key"],
                [0, vcom, com, self.l_par[0]["par"]["h"] ** sid],
            )

            f_nym.reply(
                {
                    "message": "setup",
                    "x": x,
                    "s2": s2,
                    "sig": {"R": sps_r, "S": sps_s, "T": sps_t},
                },
                record["p"],
            )

        if record["flag"] == 0:
            vcom_u = record["vcom"] * vcom

            sps_r, sps_s, sps_t = structure_preserving_signature.sign(
                self.l_par[0]["sps_secret_key"],
                [0, vcom_u, com, self.l_par[0]["par"]["h"] ** sid],
            )

            f_nym.reply(
                {
                    "message": "update",
                    "x": x,
                    "s2": s2,
                    "sig": {"R": sps_r, "S": sps_s, "T": sps_t},
                },
                record["p"],
            )


class Reader:
    """Party Rk (Reader)."""

    def __init__(self):
        pass

    l_par = []
    #   {'sid','par','par_c','com1','s1','open1'}

    l_store = []
    #   {'sid','par','par_c','sps_public_key','vcom','x','r','com','s'
    # ,'open','sig'}

    def prepare_committed_record(self, i_list):
        """Returns a list consisting of commitments to the positions in
        i_list and their corresponding values in the database x."""
        com_list = []
        for i in i_list:
            ccom_i = pedersen_commitment.commit(self.l_par[0]["par_c"], i)
            ccom_v = pedersen_commitment.commit(
                self.l_par[0]["par_c"], self.l_store[0]["x"][i]
            )
            temp_record = CommittedRecord(
                i,
                self.l_store[0]["x"][i],
                ccom_i["com"],
                ccom_i["open"],
                ccom_v["com"],
                ccom_v["open"],
            )
            com_list.append(temp_record)
        return com_list

    def update_table(self, i_list):
        """Update x with values from i_list."""
        for i in range(1, len(i_list)):
            self.l_store[0]["x"][i] = self.l_store[0]["x"][i] + i_list[i]

    def prepare_blinded_witness(self, instance, witness):
        """Blinds witnesses as required by the ZK compiler referenced
        in our paper (ref [10])."""
        # Unpack signatures
        sig = witness["sig"]
        r, s, t = sig["R"], sig["S"], sig["T"]
        sps_public_key = instance["sps_public_key"]

        # Bases for blinding elements in g and gt
        h = instance["par"]["group"].random(G1)
        ht = instance["par"]["group"].random(G2)

        [d_1, d_2, d_5] = group.random(ZR, 3)
        d_3, d_4 = witness["r2"], witness["open2"]

        # Blind sigs
        r_d = r * (h ** d_1)
        s_d = s * (h ** d_2)
        t_d = t * (ht ** d_5)

        com_d, vcom_d = instance["comd"], instance["vcomd"]

        # We refer to instance and witness values related to database
        # entries as subinstances and subwitnesses.
        subinstance_list = instance["subinstance"]
        subwitness_list = witness["subwitness"]

        index = 0
        blinded_witness = ZKWitness()
        blinded_witness.set_d(d_1, d_2, d_3, d_4, d_5)

        blinded_instance = ZKInstance()
        blinded_instance.set_bsig(r_d, s_d, t_d)
        blinded_instance.set_comd(com_d)
        blinded_instance.set_par_c(instance["par_c"])
        blinded_instance.set_par(instance["par"])
        blinded_instance.set_pk(sps_public_key)
        blinded_instance.set_sid(sid)
        blinded_instance.set_vcomd(vcom_d)
        blinded_instance.set_bh(h)
        blinded_instance.set_bht(ht)

        for subinstance_record in subinstance_list:
            subwitness_record = subwitness_list[index]

            # Unpack sigs
            sig_i = subwitness_record.sig_i
            r_i, s_i, t_i = sig_i["R"], sig_i["S"], sig_i["T"]

            [d_i_1, d_i_2, d_i_3, d_i_4, d_i_5] = group.random(ZR, 5)

            # Blind sigs
            r_i_d = r_i * (h ** d_i_1)
            s_i_d = s_i * (h ** d_i_2)
            t_i_d = t_i * (ht ** d_i_3)

            # This is par_h[i] from VC par
            par_h_i = instance["par"]["par_h"][subwitness_record.i]

            par_h_i_d = par_h_i * (ht ** d_i_4)

            # VC Witness
            w_i = subwitness_record.witness
            w_i_d = w_i * (h ** d_i_5)

            temp_blinded_subwitness_record = SubWitnessRecord(
                index,
                subwitness_record.i,
                subwitness_record.vr,
                subwitness_record.copen_i,
                subwitness_record.copen_ri,
                d_i_1,
                d_i_2,
                d_i_3,
                d_i_4,
                d_i_5,
            )
            blinded_witness.append_subwitnesses(temp_blinded_subwitness_record)
            temp_blinded_instance_record = SubInstanceRecord(
                index,
                subinstance_record.ccom_i,
                subinstance_record.ccom_ri,
                {"R_id": r_i_d, "S_id": s_i_d, "T_id": t_i_d},
                par_h_i_d,
                w_i_d,
            )
            blinded_instance.append_subinstances(temp_blinded_instance_record)
            index = index + 1
        return (
            dict_from_class(blinded_instance),
            dict_from_class(blinded_witness),
        )

    def test_witness_update(self, i_list):
        """Measures witness update times for HD."""
        if len(self.l_par) == 0 or len(self.l_store) == 0:
            print(
                "Abort: (Reader) Party hasn't been initialised."
                + " (HD witness update tests)"
            )
            exit()

        r_2 = self.l_par[0]["par"]["group"].random(ZR)
        r_d = self.l_store[0]["r"] + r_2

        for i_list_item in i_list:
            if self.l_store[0]["x"][i_list_item.i] != i_list_item.vr:
                print(
                    "Abort: (Reader) input tuples do not match HD state."
                    + " (HD witness update tests)"
                )
                exit()

            witness_1 = vector_commitment.generate_witness(
                self.l_par[0]["par"], i_list_item.i, self.l_store[0]["x"], r_d
            )

            global t_one_entry_update
            (witness_2, t_one_entry_update) = vector_commitment.update_witness(
                self.l_par[0]["par"],
                witness_1,
                i_list_item.i + 1,
                i_list_item.i,
                self.l_store[0]["x"][i_list_item.i],
                self.l_store[0]["x"][i_list_item.i] + 2,
            )

    def read(self, sid, p, i_list):
        """Read interface."""
        if len(self.l_par) == 0 or len(self.l_store) == 0:
            print("Abort: (Reader) Party hasn't been initialised. (read)")
            exit()

        r_2 = self.l_par[0]["par"]["group"].random(ZR)
        r_d = self.l_store[0]["r"] + r_2
        vcom_d = vector_commitment.rerand_com(
            self.l_par[0]["par"], self.l_store[0]["vcom"], r_2
        )

        witness_records = []
        instance_records = []
        for i_list_item in i_list:
            if self.l_store[0]["x"][i_list_item.i] != i_list_item.vr:
                print(
                    "Abort: (Reader) input tuples do not match HD"
                    + " state. (read)"
                )
                exit()
            t_comp_vcom_start = time.time()
            temp_witness_record = WitnessRecord(
                i_list_item.i,
                i_list_item.vr,
                i_list_item.copen_i,
                i_list_item.copen_ri,
                vector_commitment.generate_witness(
                    self.l_par[0]["par"],
                    i_list_item.i,
                    self.l_store[0]["x"],
                    r_d,
                ),
                self.l_store[0]["par_sig"][i_list_item.i],
            )

            global t_comp_vcom
            t_comp_vcom = str(time.time() - t_comp_vcom_start)
            witness_records.append(temp_witness_record)
            temp_instance_record = InstanceRecord(
                i_list_item.ccom_i, i_list_item.ccom_ri
            )
            instance_records.append(temp_instance_record)

        open_2 = self.l_par[0]["par"]["group"].random(ZR)
        open_d = self.l_store[0]["open"] + open_2

        com_d = pedersen_commitment.rerand(
            self.l_par[0]["par_c"], self.l_store[0]["com"], open_2
        )

        witness_read = {
            "sig": self.l_store[0]["sig"],
            "vcom": self.l_store[0]["vcom"],
            "com": self.l_store[0]["com"],
            "r2": r_2,
            "open2": open_2,
            "subwitness": witness_records,
        }
        instance_read = {
            "sid": self.l_store[0]["sid"],
            "sps_public_key": self.l_store[0]["sps_public_key"],
            "par": self.l_par[0]["par"],
            "par_c": self.l_par[0]["par_c"],
            "vcomd": vcom_d,
            "comd": com_d,
            "subinstance": instance_records,
        }
        self.l_store[0]["vcom"] = vcom_d
        self.l_store[0]["r"] = r_d
        self.l_store[0]["open"] = open_d
        p = random.randint(0, 99999999)
        self.obj_id = weak_reference.remember(self)
        blinded_instance, blinded_witness = self.prepare_blinded_witness(
            instance_read, witness_read
        )
        f_zk.prove(
            sid, blinded_witness, blinded_instance, p, self.obj_id, updater_id
        )
        return p

    def first_read(self, sid, pseudonym):
        """Read interface, first execution."""
        if len(self.l_par) == 0:
            par, par_c = f_crs.get()
            s_1 = par["group"].random(ZR)
            s_1_com = pedersen_commitment.commit_0(par_c, s_1)
            self.l_par.append(
                {
                    "sid": sid,
                    "par": par,
                    "par_c": par_c,
                    "com1": s_1_com["com"],
                    "s1": s_1,
                    "open1": s_1_com["open"],
                }
            )
            f_nym.send(
                sid,
                {"com1": s_1_com["com"], "message": "com1"},
                pseudonym,
                weak_reference.remember(self),
                updater_id,
            )
            return pseudonym

    def message_in(self, m, pseudonym):
        """Handles messages from other parties to Rk."""
        if m["message"] == "setup":
            vcom = vector_commitment.commit(self.l_par[0]["par"], m["x"], 0)
            com_2 = pedersen_commitment.commit_0(
                self.l_par[0]["par_c"], m["s2"]
            )
            com = self.l_par[0]["com1"] * com_2["com"]
            keys = f_reg.retrieve()
            sps_public_key = keys["sps_public_key"]
            par_sig = keys["par_sig"]
            if (
                structure_preserving_signature.verify(
                    sps_public_key,
                    m["sig"],
                    [
                        0,
                        vcom,
                        com,
                        self.l_par[0]["par"]["h"] ** self.l_par[0]["sid"],
                    ],
                )
                == 1
            ):
                self.l_store.append(
                    {
                        "sid": self.l_par[0]["sid"],
                        "par": self.l_par[0]["par"],
                        "par_c": self.l_par[0]["par_c"],
                        "sps_public_key": sps_public_key,
                        "vcom": vcom,
                        "x": m["x"],
                        "r": 0,
                        "com": com,
                        "s": self.l_par[0]["s1"] + m["s2"],
                        "open": self.l_par[0]["open1"],
                        "sig": m["sig"],
                        "par_sig": par_sig,
                    }
                )
            else:
                print(
                    "Abort: (Reader) Invalid SPS signatures in the"
                    + " setup phase."
                )
                exit()

        if m["message"] == "open":
            s_1 = self.l_par[0]["par"]["group"].random(ZR)
            com_1 = pedersen_commitment.commit(self.l_par[0]["par_c"], s_1)
            old_open = self.l_store[0]["open"]
            old_s_1 = self.l_store[0]["s"]

            self.l_par[0]["s1"] = s_1
            self.l_par[0]["com1"] = com_1["com"]
            self.l_par[0]["open1"] = com_1["open"]

            f_nym.send(
                self.l_store[0]["sid"],
                {
                    "message": "repcom",
                    "s1": old_s_1,
                    "open1": old_open,
                    "com1": com_1["com"],
                },
                pseudonym,
                weak_reference.remember(self),
                updater_id,
            )

        elif m["message"] == "update":
            com_2 = pedersen_commitment.commit_0(
                self.l_par[0]["par_c"], m["s2"]
            )
            com_u = self.l_par[0]["com1"] * com_2["com"]

            vcom_2 = vector_commitment.commit(self.l_par[0]["par"], m["x"], 0)
            vcom_u = self.l_store[0]["vcom"] * vcom_2

            if (
                structure_preserving_signature.verify(
                    self.l_store[0]["sps_public_key"],
                    m["sig"],
                    [
                        0,
                        vcom_u,
                        com_u,
                        self.l_par[0]["par"]["h"] ** self.l_store[0]["sid"],
                    ],
                )
                != 1
            ):
                print(
                    "Abort: (Reader) Invalid SPS signatures in the"
                    + " update phase."
                )

            self.l_store[0]["vcom"] = vcom_u
            self.l_store[0]["com"] = com_u
            self.l_store[0]["sig"] = m["sig"]
            self.update_table(m["x"])
            self.l_store[0]["s"] = self.l_par[0]["s1"] + m["s2"]
            self.l_store[0]["open"] = self.l_par[0]["open1"]

            if args.verbose:
                db_state = "; x = " + str(self.l_store[0]["x"])
            else:
                db_state = "; x"
            print(
                "Reader: hd.update.end, sid = "
                + str(self.l_par[0]["sid"])
                + "; p = "
                + str(pseudonym)
                + db_state
            )


def register():
    """PPLS Registration interface."""
    t_register_start = time.time()
    p = reader_k.first_read(sid, 0)
    for i in range(0, db_size):
        if args.randomise is True:
            db_list.append(random.randint(0, 99))
        else:
            db_list.append(0)
        empty_db_list.append(0)
    t_first_update_start = time.time()
    updater.update(sid, p, db_list)
    global t_first_update
    t_first_update = str(time.time() - t_first_update_start)
    global t_register
    t_register = str(time.time() - t_register_start)
    print(
        "Vendor: lp.register.end; sid = " + str(sid) + "; P = " + str(p) + "\n"
    )


def purchase(i, v, v_n):
    """PPLS Purchase interface."""
    t_purchase_start = time.time()
    p = random.randint(0, 99999999)
    com_list = reader_k.prepare_committed_record([2, 3])
    p = reader_k.read(sid, p, com_list)
    temp_db_list = list(empty_db_list)
    temp_db_list[i] = v
    temp_db_list[db_size - 1] = v_n
    updater.update(sid, p, temp_db_list)
    global t_purchase
    t_purchase = str(time.time() - t_purchase_start)
    print(
        "Vendor: lp.purchase.end; sid = " + str(sid) + "; P = " + str(p) + "\n"
    )


def redeem(points):
    """PPLS Redemption interface."""
    t_redeem_start = time.time()
    if points > reader_k.l_store[0]["x"][db_size]:
        print(
            "Abort: (Reader) Insufficient loyalty points for this"
            + " operation. (redeem)"
        )
    else:
        com_db_size = pedersen_commitment.commit(
            reader_k.l_par[0]["par_c"], db_size
        )
        com_v_n = pedersen_commitment.commit(
            reader_k.l_par[0]["par_c"], reader_k.l_store[0]["x"][db_size]
        )
        p = random.randint(0, 99999999)
        witness_rd = {
            "Vn": reader_k.l_store[0]["x"][db_size],
            "openVn": com_v_n["open"],
        }
        instance_rd = {
            "points": points,
            "comVn": com_v_n["com"],
            "comN": com_db_size["com"],
            "N": db_size,
            "openN": com_db_size["open"],
        }
        f_zkrd = FZK_RD(f_nym, keylength)
        f_zkrd.prove(
            sid,
            witness_rd,
            instance_rd,
            weak_reference.remember(reader_k),
            reader_k.l_par[0]["par_c"],
            group,
        )
        if (
            pedersen_commitment.verify(
                reader_k.l_par[0]["par_c"],
                com_v_n["com"],
                reader_k.l_store[0]["x"][db_size],
                com_v_n["open"],
            )
            == 0
        ):
            print(
                "Abort: (Updater) Loyalty point commitments do"
                + " not hold. (redeem)"
            )
            exit()

        com_list = reader_k.prepare_committed_record([db_size])
        p = reader_k.read(sid, p, com_list)

        temp_db_list = list(empty_db_list)
        temp_db_list[db_size - 1] = -points
        updater.update(sid, p, temp_db_list)
        global t_redeem
        t_redeem = str(time.time() - t_redeem_start)
        print(
            "Vendor: lp.redeem.end; sid = "
            + str(sid)
            + "; P = "
            + str(p)
            + "; p = "
            + str(points)
            + "\n"
        )


def profile(start, end, val):
    """PPLS Profile interface (Checks whether the sum of the values
    contained in the database between positions 'start' and 'end'
    is greater than 'val')."""
    com_list = []
    open_list = []

    for i in range(start, end + 1):
        com_position = pedersen_commitment.commit(
            reader_k.l_par[0]["par_c"], i
        )
        com_value = pedersen_commitment.commit(
            reader_k.l_par[0]["par_c"], reader_k.l_store[0]["x"][i + 1]
        )
        com_list.append(
            {"i": i, "comi": com_position["com"], "comv": com_value["com"]}
        )
        open_list.append(
            {
                "i": i,
                "v": reader_k.l_store[0]["x"][i + 1],
                "openi": com_position["open"],
                "openv": com_value["open"],
            }
        )

    witness_pr = open_list
    instance_pr = com_list

    f_zk_profile = FZK_PR3(f_nym, keylength)

    commitments, result = f_zk_profile.prove(
        sid,
        witness_pr,
        instance_pr,
        weak_reference.remember(reader_k),
        start,
        end,
        reader_k.l_par[0]["par_c"],
        group,
        val,
    )
    pseudonym = random.randint(0, 99999999)
    for i in range(start, end + 1):
        com_list = reader_k.prepare_committed_record([i])
        pseudonym = reader_k.read(sid, pseudonym, com_list)
        updater.update(sid, pseudonym, empty_db_list)

    print(
        "Vendor: lp.profile.end; sid = "
        + str(sid)
        + "; P = "
        + str(pseudonym)
        + "; res = "
        + str(result)
        + "\n"
    )


parser = argparse.ArgumentParser(
    description="Implementation of the protocol described in the"
    + " titled 'Unlinkable Updatable Hiding Databases and"
    + " Privacy-Preserving Loyalty Programs'"
)
parser.add_argument("size", metavar="N", type=int, help="Database size")

parser.add_argument(
    "-k",
    "--keylength",
    metavar="K",
    type=int,
    help="Paillier Encryption key size."
    + " (Supported values: 1024, 2048; Default: 2048)",
)
parser.add_argument(
    "-r",
    "--randomise",
    action="store_true",
    default=False,
    help="Randomise database state",
)
parser.add_argument(
    "-v",
    "--verbose",
    action="store_true",
    default=False,
    help="Display database contents and commitment values",
)

db_size = 100
keylength = 2048

args = parser.parse_args()

if args.keylength == 1024 or args.keylength == 2048:
    keylength = args.keylength

if args.size is not None and 10 < int(args.size) < 800000:
    db_size = int(args.size)
else:
    print(
        "Please enter a database size between 11 and 800000 (We need a"
        + " database containing atleast 10 entries to test the"
        + " profiling phase)."
    )
    exit()


# Curve specification
group = PairingGroup("BN256")

# Primitive instantiation
vector_commitment = VectorCommitment(group)
structure_preserving_signature = StructurePreservingSignature()
pedersen_commitment = PedersenCommitment(group)

# Functionality instantiation
f_crs = FCRS()
weak_reference = WeakReference()
f_nym = FNYM(weak_reference)
f_reg = FREG()
f_zk = FZK(f_nym, keylength)

# Updater instantiation
updater = Updater()

# Returns a unique identifer for U
updater_id = weak_reference.remember(updater)


# Init reader and set sid
reader_k = Reader()
sid = random.randint(0, 99999999)

# CRS setup
setup(db_size)

db_list = []
empty_db_list = []

# PPLS Tests
register()

purchase(1, 2, 3)

redeem(1)

# Single entry profiling
t_profile_1_start = time.time()
profile(1, 1, 20)
t_profile_1 = str(time.time() - t_profile_1_start)

# 5 entry profiling
t_profile_5_start = time.time()
profile(1, 5, 20)
t_profile_5 = str(time.time() - t_profile_5_start)

# 10 entry profiling
t_profile_10_start = time.time()
profile(1, 10, 20)
t_profile_10 = str(time.time() - t_profile_10_start)

# Uncomment for storage cost measurements
# print("CRS (par_g) in bytes = " + \
# sys.getsizeof(reader_k.l_par[0]['par']['par_g']))
# print("CRS (par_h) in bytes = " + \
# sys.getsizeof(reader_k.l_par[0]['par']['par_h']))
# print("Reader DB in bytes = " + str(get_real_size(reader_k.l_store[0]['x'])))
# print("VC par in bytes = " + str(get_real_size(reader_k.l_par[0]['par'])))
# print("Vcom in bytes = " + str(get_real_size(reader_k.l_store[0]['vcom'])))
# print("Opening in bytes = " + str(sys.getsizeof(reader_k.lwit [0])))

# HD single entry read
com_list = reader_k.prepare_committed_record([5])
t_one_entry_read_start = time.time()
p = reader_k.read(sid, 0, com_list)
t_one_entry_read = str(time.time() - t_one_entry_read_start)


temp_db_list = empty_db_list
temp_db_list[4] = temp_db_list[4] + 2
updater.update(sid, p, temp_db_list)

# HD five entry read
com_list = reader_k.prepare_committed_record([4, 5, 6, 7, 8])
t_five_entry_read_start = time.time()
reader_k.read(sid, 0, com_list)
t_five_entry_read = str(time.time() - t_five_entry_read_start)

com_list = reader_k.prepare_committed_record([5])
reader_k.test_witness_update(com_list)

output_headings = [
    "N",
    "DB Size",
    "Paillier Key Length",
    "First Update",
    "Computation of Vcom",
    "1 Entry Update",
    "1 Entry Read",
    "5 Entry Read",
    "Registration",
    "Purchase",
    "Redemption",
    "1 Entry Profiling",
    "5 Entry Profiling",
    "10 Entry Profiling",
    "Setup",
]

file_name = "UUHD-PPLS-Timing-data.xlsx"
if not Path(file_name).exists():
    results_workbook = Workbook()
    results_counter = 1
    results_max_row = 0
else:
    results_workbook = load_workbook(file_name)
    results_max_row = results_workbook.active.max_row
    results_counter = (
        results_workbook.active["A" + str(results_max_row)].value + 1
    )
results_worksheet = results_workbook.active


if results_max_row == 0:
    results_worksheet.append(output_headings)
timing_data = [
    results_counter,
    db_size,
    keylength,
    float(t_first_update),
    float(t_comp_vcom),
    float(t_one_entry_update),
    float(t_one_entry_read),
    float(t_five_entry_read),
    float(t_register),
    float(t_purchase),
    float(t_redeem),
    float(t_profile_1),
    float(t_profile_5),
    float(t_profile_10),
    float(t_setup),
]
results_worksheet.append(timing_data)
results_workbook.save(file_name)

draw_table(output_headings[:5], timing_data[:5])
draw_table(output_headings[5:10], timing_data[5:10])
draw_table(output_headings[10:], timing_data[10:])
